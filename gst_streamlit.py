import streamlit as st
import pandas as pd
import re
import io
import json
import hashlib
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import openpyxl
import base64
import gzip
import csv
import logging
import traceback
from typing import Tuple, Dict, List, Optional, Any

# ============================================================
#  LOGGING SETUP — Comprehensive Error Tracking
# ============================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ============================================================
#  EMBEDDED STATIC FILES (guide PDF + sample Excel)
# ============================================================
_GUIDE_PDF_B64   = """JVBERi0xLjQKJZOMi54gUmVwb3J0TGFiIEdlbmVyYXRlZCBQREYgZG9jdW1lbnQgKG9wZW5zb3VyY2UpCjEgMCBvYmoKPDwKL0YxIDIgMCBSIC9GMiAzIDAgUiAvRjMgNSAwIFIgL0Y0IDEwIDAgUgo+PgplbmRvYmoKMiAwIG9iag[...]"""
_SAMPLE_XL_B64   = """UEsDBBQAAAAIAI1yclxGx01IlQAAAM0AAAAQAAAAZG9jUHJvcHMvYXBwLnhtbE3PTQvCMAwG4L9SdreZih6kDkQ9ip68zy51hbYpbYT67+0EP255ecgboi6JIia2mEXxLuRtMzLHDUDWI/o+y8qhiqHke64x3YGMsRoPpB8eA8Oibd[...]"""

_HSN_HTML_GZ = "H4sIAAAAAAAC/71XwXLbNhC9+ys2OPRUipLq2JmUZBvLTZqOnWQs9dDeIBASkUAAC4CW2R/pB/XHugBJSdZQsjXTsWYkkbvAwy7wsA9IXl1/nsz++PILFG4ls7PE/4GkapkSrog3cJpnZ4CfZMUdBVZQY7lLye+z99EbsutSdMVTci/4utTG[...]"
_HSN_CSS_GZ = "H4sIAAAAAAAC/61ZWW/jthZ+z69gZ9A2HliuJC/x8tKZxG77UBRoBijuIy3RNm+0gaKzzGD+e8/hIlEy7eT2NkESmutZv/ORWYqylOTrFYGvINjug5zyYkneb9TXqu1PqEh9/QdGUyZwJNpMNws7UgmeU/EC/fF0Nl5/6vUH+6x8WhKx39Lr8[...]"
_HSN_JS1_GZ = "H4sIAAAAAAAC/+y9CXPb2JIu+FcQN6KjpQjKDYKLBDkmXoAkRKIEArwAIVmeF/GGlmibXZLoJqVbXW9i3m+fLw9AitRiJ4kEDlzlu1jc8zt7nly+vJ7fLx+MQRz8r54zdoz/y/i//99/XM9vpv84Nf5h1v9RM/7x8Oc39Qyfoac30+X1YvbtY[...]"
_HSN_JS2_GZ = "H4sIAAAAAAAC/6VZe2/bOBL/P5+C6zMqeZsoLg6L2z3HKdok7RZIdm9j3+GAplczEm3zoteRVBJf6u++M6OHKVmym9ZAkJicGc6bv2GCxM8iERuPB8HFPfxxKbURsVCuc/771VkSG1xLeCAC55C5AzY+ZU8HDD7Hx+ycG86kZiHts7lKIhbAm[...]"

def _guide_pdf_bytes() -> bytes:
    """Safely load guide PDF with fallback."""
    try:
        local_path = "c:/Users/Admin/OneDrive/Desktop/GST RECO/GST_Tool_User_Guide.pdf"
        if os.path.exists(local_path):
            with open(local_path, "rb") as f:
                return f.read()
    except Exception as e:
        logger.warning(f"Could not load local PDF: {e}")
    try:
        return base64.b64decode(_GUIDE_PDF_B64)
    except Exception as e:
        logger.error(f"Could not decode embedded PDF: {e}")
        return b""

def _sample_excel_bytes() -> bytes:
    """Safely decode sample Excel with error handling."""
    try:
        return base64.b64decode(_SAMPLE_XL_B64)
    except Exception as e:
        logger.error(f"Could not decode sample Excel: {e}")
        return b""

def _dgz(b64_str: str) -> str:
    """Safely decompress gzip-encoded base64 strings."""
    try:
        return gzip.decompress(base64.b64decode(b64_str)).decode("utf-8")
    except Exception as e:
        logger.error(f"Decompression failed: {e}")
        return ""

# ============================================================
#  PAGE CONFIG
# ============================================================
st.set_page_config(
    page_title="GST 2B Reconciliation",
    page_icon="🧾",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ============================================================
#  CSS
# ============================================================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.stApp { background: #f0f4f8; }

[data-testid="stSidebar"] { background: #1e293b !important; border-right: none !important; }
[data-testid="stSidebar"] * { color: #cbd5e1 !important; }
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3,
[data-testid="stSidebar"] h4 { color: #f1f5f9 !important; font-weight: 700 !important; }
[data-testid="stSidebar"] .stMarkdown p { color: #94a3b8 !important; font-size: 0.82rem !important; }
[data-testid="stSidebar"] hr { border-color: #334155 !important; }

/* ── Top header ── */
.top-header {
    background: linear-gradient(135deg, #1e3a5f 0%, #1d4ed8 60%, #1e40af 100%);
    border-radius: 14px; padding: 28px 36px; margin-bottom: 24px;
    display: flex; align-items: center; justify-content: space-between;
    box-shadow: 0 4px 20px rgba(29,78,216,0.2);
}
.header-left  { display: flex; align-items: center; gap: 16px; }
.header-icon  { width: 52px; height: 52px; background: rgba(255,255,255,0.15);
                border-radius: 12px; display: flex; align-items: center;
                justify-content: center; font-size: 1.6rem; border: 1px solid rgba(255,255,255,0.2); }
.header-title { font-size: 1.5rem; font-weight: 800; color: #ffffff; letter-spacing: -0.5px; margin: 0; }
.header-sub   { font-size: 0.82rem; color: rgba(255,255,255,0.7); margin: 3px 0 0 0; }
.header-badge { background: rgba(255,255,255,0.15); border: 1px solid rgba(255,255,255,0.25);
                color: #ffffff; font-size: 0.72rem; font-weight: 600; padding: 6px 14px;
                border-radius: 20px; letter-spacing: 0.5px; }

/* ── Labels ── */
.box-label    { font-size: 0.75rem; font-weight: 700; color: #374151;
                text-transform: uppercase; letter-spacing: 0.8px; margin-bottom: 10px; display: block; }
.section-label{ font-size: 0.7rem; font-weight: 700; color: #2563eb;
                text-transform: uppercase; letter-spacing: 1.2px; margin-bottom: 4px; }
.section-title{ font-size: 1.15rem; font-weight: 700; color: #111827;
                margin-bottom: 16px; letter-spacing: -0.3px; }

/* ── Legend ── */
.legend-row { display: flex; gap: 10px; flex-wrap: wrap; margin: 12px 0 18px 0; }
.leg        { display: inline-flex; align-items: center; gap: 6px; padding: 5px 12px;
              border-radius: 20px; font-size: 0.76rem; font-weight: 600; }
.leg-green  { background: #dcfce7; color: #166534; border: 1px solid #bbf7d0; }
.leg-yellow { background: #fef9c3; color: #854d0e; border: 1px solid #fde68a; }
.leg-red    { background: #fee2e2; color: #991b1b; border: 1px solid #fecaca; }
.leg-blue   { background: #dbeafe; color: #1e40af; border: 1px solid #bfdbfe; }

/* ── Metrics ── */
[data-testid="metric-container"] {
    background: #ffffff !important; border: 1px solid #e5e7eb !important;
    border-radius: 12px !important; padding: 18px !important;
    box-shadow: 0 1px 6px rgba(0,0,0,0.05) !important;
}

/* ── Buttons ── */
.stButton > button {
    background: linear-gradient(135deg, #1d4ed8, #2563eb) !important;
    color: #ffffff !important; border: none !important; border-radius: 10px !important;
    font-weight: 700 !important; font-size: 0.95rem !important;
    padding: 13px 28px !important; width: 100% !important;
    box-shadow: 0 4px 14px rgba(37,99,235,0.3) !important;
}

hr { border-color: #e2e8f0 !important; margin: 24px 0 !important; }

footer { visibility: hidden !important; display: none !important; }
#MainMenu { visibility: hidden !important; display: none !important; }
header[data-testid="stHeader"] { display: none !important; }
a[href*="github.com"] { display: none !important; }
[data-testid="stToolbar"] { display: none !important; }
</style>
""", unsafe_allow_html=True)

# ============================================================
#  USER / USAGE MANAGEMENT
# ============================================================
FREE_LIMIT = 5
PRICE_INR  = 79
DATA_FILE  = "gst_users.json"

def _load_users() -> dict:
    """Safely load user database."""
    try:
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception as e:
        logger.warning(f"Failed to load users: {e}")
    return {}

def _save_users(db: dict) -> None:
    """Safely save user database."""
    try:
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(db, f, indent=2)
    except Exception as e:
        logger.warning(f"Failed to save users: {e}")

def _year_key() -> str:
    """Get year key."""
    return datetime.now().strftime("%Y")

def _month_key() -> str:
    """Get month key."""
    return datetime.now().strftime("%Y-%m")

def _uid(email: str) -> str:
    """Generate user ID from email."""
    try:
        return hashlib.md5(email.strip().lower().encode()).hexdigest()
    except Exception as e:
        logger.error(f"UID generation failed: {e}")
        return ""

def _get_user(email: str) -> dict:
    """Get or create user record."""
    try:
        db = _load_users()
        uid = _uid(email)
        if uid not in db:
            db[uid] = {
                "email": email.strip().lower(),
                "name": "",
                "firm": "",
                "gstin": "",
                "registered": datetime.now().strftime("%d-%b-%Y"),
                "paid_until": "",
                "lifetime_usage": 0,
                "usage": {}
            }
            _save_users(db)
        return db[uid]
    except Exception as e:
        logger.error(f"Get user failed: {e}")
        return {}

def _get_yearly_usage(email: str) -> int:
    """Get yearly usage count."""
    try:
        return _get_user(email).get("lifetime_usage", 0)
    except Exception:
        return 0

def _increment_usage(email: str) -> None:
    """Increment usage counter."""
    try:
        db = _load_users()
        uid = _uid(email)
        if uid in db:
            db[uid]["lifetime_usage"] = db[uid].get("lifetime_usage", 0) + 1
            _save_users(db)
    except Exception as e:
        logger.warning(f"Increment usage failed: {e}")

def _is_paid(email: str) -> bool:
    """Check if user is paid."""
    try:
        return _get_user(email).get("paid_until", "") >= _month_key()
    except Exception:
        return False

def _mark_paid(email: str) -> None:
    """Mark user as paid."""
    try:
        db = _load_users()
        uid = _uid(email)
        if uid in db:
            db[uid]["paid_until"] = _month_key()
            _save_users(db)
    except Exception as e:
        logger.warning(f"Mark paid failed: {e}")

def _update_profile(email: str, name: str, firm: str, gstin: str) -> None:
    """Update user profile."""
    try:
        db = _load_users()
        uid = _uid(email)
        if uid in db:
            db[uid].update({"name": name, "firm": firm, "gstin": gstin})
            _save_users(db)
    except Exception as e:
        logger.warning(f"Update profile failed: {e}")

def _can_run(email: str) -> Tuple[bool, str]:
    """Check if user can run reconciliation."""
    try:
        if _is_paid(email):
            return True, "paid"
        used = _get_yearly_usage(email)
        if used < FREE_LIMIT:
            return True, "free"
        return False, "limit_exceeded"
    except Exception:
        return False, "error"

# ============================================================
#  GSTR-2B COLUMN KEYWORDS
# ============================================================
GSTR2B_COL_KEYWORDS = {
    "GSTIN": ["gstin of supplier", "gstin/uin of supplier", "gstin", "gst no"],
    "Vendor": ["trade/legal name", "trade name", "legal name", "supplier name"],
    "Invoice No": ["invoice number", "invoice no.", "invoice no", "inv no"],
    "Invoice Type": ["invoice type", "document type", "type of invoice"],
    "Invoice Date": ["invoice date", "bill date", "document date"],
    "Invoice Value": ["invoice value", "bill value", "total invoice value"],
    "Taxable": ["taxable value", "taxable amount", "value of supply"],
    "IGST": ["integrated tax", "igst amount", "igst"],
    "CGST": ["central tax", "cgst amount", "cgst"],
    "SGST": ["state/ut tax", "sgst amount", "sgst"],
}

B2B_COL_POS_FALLBACK = {
    1: "GSTIN", 2: "Vendor", 3: "Invoice No", 4: "Invoice Type",
    5: "Invoice Date", 6: "Invoice Value", 7: "Taxable", 8: "IGST",
    9: "CGST", 10: "SGST", 11: "Cess",
}

BOOKS_COL_MAP = {
    "GSTIN": ["gstin of supplier", "supplier gstin", "party gstin", "gstin"],
    "Vendor": ["trade/legal name", "supplier name", "party name", "vendor name"],
    "Invoice No": ["invoice number", "invoice no", "bill number", "bill no"],
    "Invoice Date": ["invoice date", "bill date", "voucher date"],
    "Invoice Value": ["invoice value", "bill value", "total invoice value"],
    "Taxable": ["taxable value", "taxable amount", "basic value"],
    "IGST": ["integrated tax", "igst amount", "igst"],
    "CGST": ["central tax", "cgst amount", "cgst"],
    "SGST": ["state/ut tax", "sgst amount", "sgst"],
}

CDNR_BOOKS_COL_MAP = {
    "GSTIN": ["gstin of supplier", "supplier gstin", "gstin"],
    "Vendor": ["trade/legal name", "supplier name", "party name"],
    "Note No": ["credit note number", "debit note number", "note number"],
    "Note Date": ["credit note date", "debit note date", "note date"],
    "Note Type": ["note type", "credit/debit", "type of note"],
    "Note Value": ["credit note value", "debit note value", "note value"],
    "Taxable": ["taxable value", "taxable amount"],
    "IGST": ["integrated tax", "igst"],
    "CGST": ["central tax", "cgst"],
    "SGST": ["state/ut tax", "sgst"],
}

# ============================================================
#  COLUMN FINDER
# ============================================================
def _clean(s: str) -> str:
    """Clean column name for matching."""
    try:
        return re.sub(r"[^a-z0-9 ]", " ", str(s).lower()).strip()
    except Exception:
        return ""

def _nodot(s: str) -> str:
    """Remove all non-alphanumeric characters."""
    try:
        return re.sub(r"[^a-z0-9]", "", str(s).lower())
    except Exception:
        return ""

def _find_col(columns: List, field: str, kw_map: Dict) -> Optional[str]:
    """Find column by keywords."""
    try:
        keywords = kw_map.get(field, [field.lower()])
        lower_map = {c.lower().strip(): c for c in columns}
        clean_map = {_clean(c): c for c in columns}
        
        for kw in keywords:
            if kw.lower() in lower_map:
                return lower_map[kw.lower()]
        for kw in keywords:
            if _clean(kw) in clean_map:
                return clean_map[_clean(kw)]
        return None
    except Exception as e:
        logger.error(f"Find column failed: {e}")
        return None

# ============================================================
#  HEADER ROW FINDER
# ============================================================
def _find_header_row(all_rows: List, max_scan: int = 20, col_keywords: Optional[List] = None) -> int:
    """Find header row in data."""
    try:
        if col_keywords is None:
            col_keywords = ["gstin", "gst", "invoice", "bill", "taxable", "date"]
        best_idx = None
        best_score = 0
        
        for i, row in enumerate(all_rows[:max_scan]):
            row_str = " ".join(str(v).lower() for v in row if v is not None)
            score = sum(1 for kw in col_keywords if kw in row_str)
            non_none = sum(1 for v in row if v is not None)
            
            if non_none >= 3:
                if score > best_score:
                    best_score = score
                    best_idx = i
                elif best_idx is None:
                    best_idx = i
        
        return best_idx if best_idx is not None else 0
    except Exception as e:
        logger.error(f"Find header row failed: {e}")
        return 0

# ============================================================
#  GSTR-2B LOADER
# ============================================================
def _load_b2b_sheet(file_obj, sheet_type: str = "B2B") -> Tuple:
    """Load GSTR-2B sheet with error handling."""
    try:
        file_obj.seek(0)
        raw = file_obj.read()
        
        wb = None
        for _ro in [False, True]:
            try:
                wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=_ro, data_only=True)
                break
            except Exception:
                continue
        
        if wb is None:
            try:
                xf = pd.ExcelFile(io.BytesIO(raw))
                return pd.DataFrame(), "", {}
            except Exception:
                raise ValueError("Excel file is corrupted or invalid.")
        
        sheets = wb.sheetnames
        b2b_name = None
        
        if sheet_type == "CDNR":
            for s in sheets:
                if any(k in s.lower() for k in ["cdnr", "credit", "debit", "note"]):
                    b2b_name = s
                    break
            if not b2b_name:
                return pd.DataFrame(), "", {}
        else:
            for s in sheets:
                if s.strip().upper() == "B2B":
                    b2b_name = s
                    break
            if not b2b_name:
                for s in sheets:
                    if "b2b" in s.lower() and "cdnr" not in s.lower():
                        b2b_name = s
                        break
            if not b2b_name:
                raise ValueError(f"Could not find B2B sheet. Found: {sheets}")
        
        ws = wb[b2b_name]
        all_rows = list(ws.iter_rows(values_only=True))
        
        if not all_rows:
            return pd.DataFrame(columns=list(B2B_COL_POS_FALLBACK.values())), b2b_name, {}
        
        det = {"method": "", "header_row": 0, "col_map": {}}
        header_idx = _find_header_row(all_rows, max_scan=20)
        det["header_row"] = header_idx + 1
        
        hr = [str(v).strip() if v is not None else "" for v in all_rows[header_idx]]
        merged = [f"Col_{i}" if not h else h for i, h in enumerate(hr)]
        
        col_map = {}
        for field in GSTR2B_COL_KEYWORDS:
            found = _find_col(merged, field, GSTR2B_COL_KEYWORDS)
            if found and found not in col_map.values():
                col_map[field] = found
        
        essential = ["GSTIN", "Invoice No", "Taxable"]
        found_ess = sum(1 for e in essential if e in col_map)
        
        data_start = None
        for i, row in enumerate(all_rows):
            for v in row:
                sv = str(v).strip().upper().replace(" ", "")
                if len(sv) == 15 and sv[:2].isdigit():
                    data_start = i
                    break
            if data_start is not None:
                break
        
        if data_start is None:
            return pd.DataFrame(columns=list(B2B_COL_POS_FALLBACK.values())), b2b_name, det
        
        det["method"] = f"Column Detection ({found_ess}/3)"
        det["col_map"] = col_map
        
        records = []
        col_idx_map = {}
        for field, orig in col_map.items():
            for i, h in enumerate(merged):
                if h == orig:
                    col_idx_map[field] = i
                    break
        
        for row in all_rows[data_start:]:
            if not any(v is not None for v in row):
                continue
            gi = col_idx_map.get("GSTIN", 0)
            gv = row[gi] if gi < len(row) else None
            if gv is None:
                continue
            sv = str(gv).strip().upper().replace(" ", "")
            if not (8 <= len(sv) <= 15):
                continue
            
            rec = {}
            for field in B2B_COL_POS_FALLBACK.values():
                idx = col_idx_map.get(field)
                rec[field] = row[idx] if idx is not None and idx < len(row) else None
            records.append(rec)
        
        if not records:
            return pd.DataFrame(columns=list(B2B_COL_POS_FALLBACK.values())), b2b_name, det
        
        df = pd.DataFrame(records)
        df["GSTIN"] = df["GSTIN"].astype(str).str.strip().str.upper()
        df = df[df["GSTIN"].str.match(r"^\d{2}[A-Z]{5}\d{4}[A-Z]\d[Z][A-Z0-9]$", na=False)].copy()
        
        for col in ["Taxable", "IGST", "CGST", "SGST", "Cess", "Invoice Value"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
            else:
                df[col] = 0.0
        
        df["Total Tax"] = df["IGST"] + df["CGST"] + df["SGST"]
        return df.reset_index(drop=True), b2b_name, det
    
    except Exception as e:
        logger.error(f"Load B2B sheet failed: {e}")
        raise ValueError(str(e))

# ============================================================
#  BOOKS FILE READER — Multi-engine with full fallback
# ============================================================
def _try_read_raw(raw: bytes) -> List[Tuple]:
    """Try every engine to load Excel bytes."""
    results = []
    
    # Engine 1: openpyxl normal mode
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=False, data_only=True)
        for sn in wb.sheetnames:
            rows = list(wb[sn].iter_rows(values_only=True))
            if rows:
                results.append((sn, rows))
        if results:
            return results
    except Exception:
        pass
    
    # Engine 2: openpyxl read_only mode
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        for sn in wb.sheetnames:
            rows = list(wb[sn].iter_rows(values_only=True))
            if rows:
                results.append((sn, rows))
        wb.close()
        if results:
            return results
    except Exception:
        pass
    
    # Engine 3: pandas openpyxl
    try:
        xf = pd.ExcelFile(io.BytesIO(raw), engine="openpyxl")
        for sn in xf.sheet_names:
            df = xf.parse(sn, header=None)
            if not df.empty:
                rows = [tuple(r) for r in df.values.tolist()]
                results.append((sn, rows))
        if results:
            return results
    except Exception:
        pass
    
    # Engine 4: pandas xlrd
    try:
        xf = pd.ExcelFile(io.BytesIO(raw), engine="xlrd")
        for sn in xf.sheet_names:
            df = xf.parse(sn, header=None)
            if not df.empty:
                rows = [tuple(r) for r in df.values.tolist()]
                results.append((sn, rows))
        if results:
            return results
    except Exception:
        pass
    
    # Engine 5: pandas auto
    try:
        xf = pd.ExcelFile(io.BytesIO(raw))
        for sn in xf.sheet_names:
            df = xf.parse(sn, header=None)
            if not df.empty:
                rows = [tuple(r) for r in df.values.tolist()]
                results.append((sn, rows))
        if results:
            return results
    except Exception:
        pass
    
    # Engine 6: CSV fallback
    try:
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = [tuple(r) for r in reader if any(c.strip() for c in r)]
        if len(rows) > 1:
            return [("Sheet1", rows)]
    except Exception:
        pass
    
    raise ValueError("Could not open file. Please ensure it's a valid Excel file.")

def _read_books_file(file_obj, manual_header_row: Optional[int] = None, 
                     manual_sheet: Optional[str] = None, sheet_type: str = "B2B") -> Tuple:
    """Read books file with multi-engine support."""
    try:
        file_obj.seek(0)
        raw = file_obj.read()
        
        sheet_data = _try_read_raw(raw)
        
        best_df = None
        best_score = -1
        best_info = {}
        target = ([s for s in sheet_data if s[0] == manual_sheet] 
                  if manual_sheet else sheet_data)
        
        if not target:
            target = sheet_data
        
        CDNR_SHEET_KEYWORDS = [
            "cdnr", "cdn", "credit note", "debit note", "note", "return"
        ]
        
        if sheet_type == "CDNR":
            cdnr_target = []
            for sn, rows in target:
                sn_l = sn.lower().replace(" ", "")
                if any(kw.replace(" ", "") in sn_l for kw in CDNR_SHEET_KEYWORDS):
                    cdnr_target.append((sn, rows))
            if not cdnr_target:
                return pd.DataFrame(), {}
            target = cdnr_target
        
        for sn, all_rows in target:
            if not all_rows:
                continue
            
            hidx = (max(0, manual_header_row - 1) if manual_header_row is not None
                    else _find_header_row(all_rows, max_scan=20))
            
            if hidx >= len(all_rows):
                continue
            
            header = []
            seen = {}
            for j, v in enumerate(all_rows[hidx]):
                h = str(v).strip() if v is not None else f"Col_{j}"
                if h in seen:
                    seen[h] += 1
                    h = f"{h}_{seen[h]}"
                else:
                    seen[h] = 0
                header.append(h)
            
            data_rows = []
            for row in all_rows[hidx + 1:]:
                if any(v is not None and str(v).strip() not in ("", "None", "nan") 
                       for v in row):
                    rl = list(row)
                    while len(rl) < len(header):
                        rl.append(None)
                    data_rows.append(rl[:len(header)])
            
            if not data_rows:
                continue
            
            df = pd.DataFrame(data_rows, columns=header).dropna(how="all").reset_index(drop=True)
            score = 0
            for col in df.columns:
                vals = df[col].dropna().astype(str).str.strip()
                score += vals.str.match(r"^\d{2}[A-Z0-9]{13}$").sum() * 10
                if any(k in col.lower() for k in 
                       ["gstin", "gst", "invoice", "bill", "taxable"]):
                    score += 5
            
            info = {"sheet": sn, "header_row": hidx + 1, "score": score}
            if score > best_score:
                best_score = score
                best_df = df
                best_info = info
        
        if best_df is None or best_df.empty:
            raise ValueError("No data found in file.")
        
        return best_df, best_info
    
    except Exception as e:
        logger.error(f"Read books file failed: {e}")
        raise ValueError(str(e))

# ============================================================
#  BOOKS NORMALIZER
# ============================================================
def _load_books(df_raw, extra_hints: Optional[Dict] = None) -> Tuple:
    """Normalize and load books data."""
    try:
        df = df_raw.copy()
        df.columns = [str(c).strip() for c in df.columns]
        
        _priority = ["GSTIN", "Invoice No", "Invoice Date", "Taxable", 
                     "IGST", "CGST", "SGST", "Vendor"]
        _other = [f for f in BOOKS_COL_MAP if f not in _priority]
        mapping = {}
        _claimed = set()
        
        for field in _priority + _other:
            found = _find_col(list(df.columns), field, BOOKS_COL_MAP)
            if found and found not in _claimed:
                mapping[field] = found
                _claimed.add(found)
        
        df = df.rename(columns={v: k for k, v in mapping.items() if v in df.columns})
        
        for col in ["GSTIN", "Invoice No", "Vendor"]:
            if col not in df.columns:
                df[col] = ""
            df[col] = df[col].fillna("").astype(str).str.strip()
        
        # FIX: Use raw string to avoid escape sequence errors
        df["GSTIN"] = (df["GSTIN"]
                       .str.upper()
                       .str.replace(r"[\s\u00a0\u200b\u200c\u200d\ufeff]+", "", regex=True))
        
        gstin_15 = df["GSTIN"].str.len() == 15
        gstin_start = df["GSTIN"].str.match(r"^\d{2}", na=False)
        gstin_has_letters = df["GSTIN"].str.contains(r"[A-Z]", na=False)
        gstin_partial = df["GSTIN"].str.len().between(10, 14)
        gstin_junk = df["GSTIN"].str.match(
            r"^(NAN|NONE|TOTAL|GRAND|SUB|GSTIN|NA|NIL|N/A|NULL|-).*$", na=False)
        
        df = df[(((gstin_15 & gstin_start & gstin_has_letters) | gstin_partial) & ~gstin_junk)].copy()
        
        if df.empty:
            for c in ["IGST", "CGST", "SGST", "Taxable"]:
                df[c] = 0.0
            df["Total Tax"] = 0.0
            df["Invoice Date"] = pd.NaT
            return df.reset_index(drop=True), mapping
        
        if "Invoice Date" not in df.columns:
            df["Invoice Date"] = pd.NaT
        else:
            df["Invoice Date"] = pd.to_datetime(df["Invoice Date"], errors="coerce", dayfirst=True)
        
        for c in ["IGST", "CGST", "SGST", "Taxable"]:
            if c not in df.columns:
                df[c] = 0.0
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
        
        df["Total Tax"] = df["IGST"] + df["CGST"] + df["SGST"]
        return df.reset_index(drop=True), mapping
    
    except Exception as e:
        logger.error(f"Load books failed: {e}")
        raise ValueError(str(e))

# ============================================================
#  INVOICE MATCHING
# ============================================================
def _norm(s: str) -> str:
    """Normalize for matching."""
    try:
        return re.sub(r"[^a-z0-9]", "", str(s).lower().strip())
    except Exception:
        return ""

def _match_invoice(a: str, b: str) -> Tuple[bool, str]:
    """Match invoices with smart logic."""
    try:
        if not a or not b or str(a) in ("nan", "None", "") or str(b) in ("nan", "None", ""):
            return False, ""
        
        na, nb = _norm(a), _norm(b)
        if na == nb and na:
            return True, "Exact"
        if len(na) >= 4 and len(nb) >= 4:
            if na[-4:] == nb[-4:]:
                return True, "Suffix Match"
        return False, ""
    except Exception:
        return False, ""

# ============================================================
#  RECONCILIATION ENGINE
# ============================================================
def run_reco(gstr2b_file, books_df_raw, date_tol: float, amt_tol: float, 
             taxable_tol: float, extra_hints: Optional[Dict] = None, 
             sheet_type: str = "B2B") -> Tuple:
    """Run reconciliation with error handling."""
    try:
        gstr2b, b2b_sheet, gstr2b_det = _load_b2b_sheet(gstr2b_file, sheet_type=sheet_type)
        if gstr2b.empty:
            raise ValueError("No data found in GSTR-2B file.")
        
        books, b_map = _load_books(books_df_raw, extra_hints)
        if books.empty:
            raise ValueError("No data found in Purchase Register.")
        
        by_gstin = {}
        for j, bk in books.iterrows():
            by_gstin.setdefault(bk["GSTIN"], []).append((j, bk))
        
        results = []
        matched = set()
        
        for _, r2b in gstr2b.iterrows():
            g2b = r2b["GSTIN"]
            i2b = str(r2b.get("Invoice No", ""))
            bj = bbk = None
            bm = ""
            
            cands = list(by_gstin.get(g2b, []))
            for j, bk in cands:
                if j in matched:
                    continue
                ok, mt = _match_invoice(i2b, str(bk.get("Invoice No", "")))
                if ok:
                    bj, bbk, bm = j, bk, mt
                    break
            
            if bbk is not None:
                matched.add(bj)
                bk = bbk
                diffs = []
                
                results.append({
                    "GSTIN": g2b,
                    "Vendor (2B)": r2b.get("Vendor", ""),
                    "Vendor (Books)": bk.get("Vendor", ""),
                    "Invoice No (2B)": i2b,
                    "Invoice No (Books)": bk.get("Invoice No", ""),
                    "Invoice Date (2B)": r2b.get("Invoice Date"),
                    "Invoice Date (Books)": bk.get("Invoice Date"),
                    "Taxable (2B)": r2b.get("Taxable", 0),
                    "Taxable (Books)": bk.get("Taxable", 0),
                    "IGST (2B)": r2b.get("IGST", 0),
                    "IGST (Books)": bk.get("IGST", 0),
                    "CGST (2B)": r2b.get("CGST", 0),
                    "CGST (Books)": bk.get("CGST", 0),
                    "SGST (2B)": r2b.get("SGST", 0),
                    "SGST (Books)": bk.get("SGST", 0),
                    "Total Tax (2B)": r2b.get("Total Tax", 0),
                    "Total Tax (Books)": bk.get("Total Tax", 0),
                    "Status": "MATCHED",
                    "Remarks": "Matched"
                })
            else:
                results.append({
                    "GSTIN": g2b,
                    "Vendor (2B)": r2b.get("Vendor", ""),
                    "Vendor (Books)": "",
                    "Invoice No (2B)": i2b,
                    "Invoice No (Books)": "",
                    "Invoice Date (2B)": r2b.get("Invoice Date"),
                    "Invoice Date (Books)": pd.NaT,
                    "Taxable (2B)": r2b.get("Taxable", 0),
                    "Taxable (Books)": None,
                    "IGST (2B)": r2b.get("IGST", 0),
                    "IGST (Books)": None,
                    "CGST (2B)": r2b.get("CGST", 0),
                    "CGST (Books)": None,
                    "SGST (2B)": r2b.get("SGST", 0),
                    "SGST (Books)": None,
                    "Total Tax (2B)": r2b.get("Total Tax", 0),
                    "Total Tax (Books)": None,
                    "Status": "IN 2B – NOT IN BOOKS",
                    "Remarks": "Not found in books"
                })
        
        for j, bk in books.iterrows():
            if j not in matched:
                results.append({
                    "GSTIN": bk["GSTIN"],
                    "Vendor (2B)": "",
                    "Vendor (Books)": bk.get("Vendor", ""),
                    "Invoice No (2B)": "",
                    "Invoice No (Books)": bk.get("Invoice No", ""),
                    "Invoice Date (2B)": pd.NaT,
                    "Invoice Date (Books)": bk.get("Invoice Date"),
                    "Taxable (2B)": None,
                    "Taxable (Books)": bk.get("Taxable", 0),
                    "IGST (2B)": None,
                    "IGST (Books)": bk.get("IGST", 0),
                    "CGST (2B)": None,
                    "CGST (Books)": bk.get("CGST", 0),
                    "SGST (2B)": None,
                    "SGST (Books)": bk.get("SGST", 0),
                    "Total Tax (2B)": None,
                    "Total Tax (Books)": bk.get("Total Tax", 0),
                    "Status": "IN BOOKS – NOT IN 2B",
                    "Remarks": "Not uploaded to portal"
                })
        
        out = pd.DataFrame(results) if results else pd.DataFrame()
        meta = {"b2b_sheet": b2b_sheet, "b2b_count": len(gstr2b), 
                "books_count": len(books), "b_map": b_map}
        return out, meta
    
    except Exception as e:
        logger.error(f"Reconciliation failed: {e}")
        raise ValueError(str(e))

# ============================================================
#  EXCEL REPORT BUILDER
# ============================================================
def build_excel(df: pd.DataFrame) -> bytes:
    """Build Excel report safely."""
    try:
        wb = Workbook()
        wb.remove(wb.active)
        thin = Side(style="thin", color="D0D0D0")
        bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        ws = wb.create_sheet("Reconciliation")
        ws.merge_cells(f"A1:Z1")
        ws["A1"].value = f"GST 2B vs Books — {datetime.now().strftime('%d-%b-%Y')}"
        ws["A1"].font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="1F3864")
        
        for ci, col in enumerate(df.columns, 1):
            c = ws.cell(row=2, column=ci, value=col)
            c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="2E75B6")
            c.border = bdr
        
        for ri, (_, row) in enumerate(df.iterrows(), 3):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = bdr
        
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()
    
    except Exception as e:
        logger.error(f"Build Excel failed: {e}")
        return b""

# ============================================================
#  SESSION STATE INIT
# ============================================================
if "ran" not in st.session_state:
    st.session_state["ran"] = False
if "logged_in_email" not in st.session_state:
    st.session_state["logged_in_email"] = ""
if "anon_runs" not in st.session_state:
    st.session_state["anon_runs"] = 0
if "current_view" not in st.session_state:
    st.session_state["current_view"] = "gst_reco"

# ============================================================
#  MAIN APP
# ============================================================
st.title("GST 2B Reconciliation Tool")

with st.sidebar:
    st.markdown("#### 📥 Help & Downloads")
    DATE_TOL = st.slider("Date Tolerance (days)", 0, 15, 5)
    AMT_TOL = st.slider("GST Tolerance (₹)", 0.0, 50.0, 2.0, step=0.5)
    TAXABLE_TOL = st.slider("Taxable Tolerance (₹)", 0.0, 100.0, 10.0, step=1.0)

st.markdown("---")
col1, col2 = st.columns(2)

with col1:
    st.markdown("**📥 GSTR-2B File**")
    gstr_file = st.file_uploader("Select GSTR-2B", type=["xlsx", "xls"], key="gstr")

with col2:
    st.markdown("**📗 Purchase Register**")
    books_file = st.file_uploader("Select Books", type=["xlsx", "xls"], key="books")

if st.button("▶️ Run Reconciliation", use_container_width=True):
    if not gstr_file or not books_file:
        st.error("Please upload both files")
    else:
        try:
            with st.spinner("Processing..."):
                books_file.seek(0)
                b_df, parse_info = _read_books_file(books_file)
                gstr_file.seek(0)
                result_df, meta = run_reco(gstr_file, b_df, DATE_TOL, AMT_TOL, TAXABLE_TOL)
                
                st.success("✅ Reconciliation complete!")
                st.dataframe(result_df, use_container_width=True)
                
                col1, col2 = st.columns(2)
                with col1:
                    st.download_button(
                        "📥 Download Report",
                        data=build_excel(result_df),
                        file_name=f"GST_Reco_{datetime.now().strftime('%d%b%Y_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        
        except Exception as e:
            logger.error(f"App error: {e}\n{traceback.format_exc()}")
            st.error(f"❌ Error: {str(e)}")

st.markdown("---")
st.markdown(f"© {datetime.now().year} GST Reconciliation Tool - All Rights Reserved")
